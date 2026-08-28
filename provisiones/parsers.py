"""
Lectura del archivo de detalle de provisiones (hoja "PROVISIONES"): un
renglón por gasto a provisionar, con sector, proveedor, cuenta, centro de
costo, código de actividad e importe.

Los importes vienen a veces como número (celda numérica de Excel) y a veces
como texto con formato argentino ("$ 1.050.537,47" o "$ 782.808" sin
decimales), así que se normalizan acá.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Optional

import openpyxl

HOJA = "PROVISIONES"
FILA_ENCABEZADO = 2  # la fila 1 es el título ("Provisiones AAAA"), la 2 trae los nombres de columna

MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}
_MES_A_NUMERO = {nombre: numero for numero, nombre in MESES_ES.items()}


@dataclass
class Provision:
    sector: str
    mes: int  # 1-12, derivado de "Periodo de aplicación"
    proveedor: str
    descripcion: str
    importe: float
    cuenta: str
    centro_costo: str
    codigo_actividad: Any  # se preserva tal cual viene (texto o número)
    referencia: str  # columna "OC - RG - CF - RC"
    comentario: Optional[str]


def _parse_importe(valor: Any) -> float:
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace("$", "").replace("\xa0", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(".", "")
    return float(texto)


def _parse_mes(valor: Any) -> int:
    nombre = str(valor).strip().lower()
    if nombre not in _MES_A_NUMERO:
        raise ValueError(f"Periodo de aplicación no reconocido: {valor!r}")
    return _MES_A_NUMERO[nombre]


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def leer_provisiones(path: str) -> list[Provision]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(f"El archivo no tiene una hoja llamada {HOJA!r} (hojas: {wb.sheetnames})")
    ws = wb[HOJA]

    provisiones = []
    for fila in ws.iter_rows(min_row=FILA_ENCABEZADO + 1, max_row=ws.max_row, values_only=True):
        sector = fila[0]
        if not sector:
            continue
        importe = _parse_importe(fila[5])
        if importe == 0:
            continue
        provisiones.append(
            Provision(
                sector=str(sector).strip(),
                mes=_parse_mes(fila[1]),
                proveedor=_texto(fila[3]),
                descripcion=_texto(fila[4]),
                importe=importe,
                cuenta=_texto(fila[6]),
                centro_costo=_texto(fila[9]),
                codigo_actividad=fila[10],
                referencia=_texto(fila[11]),
                comentario=_texto(fila[15]) or None,
            )
        )
    return provisiones
