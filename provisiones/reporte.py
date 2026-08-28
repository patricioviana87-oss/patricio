"""Genera el Excel de un asiento de provisión, con el mismo layout de columnas
que usa el sistema contable para importar asientos (ver archivo de ejemplo
en el README)."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COLUMNAS = [
    "Numero", "Fecha", "Detalle", "Fecha Aplicacion", "C.Costos", "Moneda",
    "Cotizacion", "Cuenta", "Detalle Item", "Tipo (D|H)", "Importe",
    "Cod.Ejercicio", "Codigo Actividad", "Nota",
]
ANCHOS = [8, 10, 35, 14, 9, 8, 10, 9, 55, 6, 14, 12, 14, 30]

FORMATO_FECHA = "mm-dd-yy"
FORMATO_IMPORTE = '0.00_ ;[Red]\\-0.00\\ '


@dataclass
class Linea:
    numero: int
    fecha: date
    detalle: str
    fecha_aplicacion: date
    centro_costo: str
    moneda: str
    cotizacion: float
    cuenta: str
    detalle_item: str
    tipo: str
    importe: float
    cod_ejercicio: str
    codigo_actividad: Any
    nota: Optional[str]


def escribir_asiento(lineas: list[Linea], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for j, encabezado in enumerate(COLUMNAS, start=1):
        ws.cell(row=1, column=j, value=encabezado)

    for i, linea in enumerate(lineas, start=2):
        ws.cell(row=i, column=1, value=linea.numero)
        ws.cell(row=i, column=2, value=linea.fecha).number_format = FORMATO_FECHA
        ws.cell(row=i, column=3, value=linea.detalle)
        ws.cell(row=i, column=4, value=linea.fecha_aplicacion).number_format = FORMATO_FECHA
        ws.cell(row=i, column=5, value=str(linea.centro_costo))
        ws.cell(row=i, column=6, value=linea.moneda)
        ws.cell(row=i, column=7, value=linea.cotizacion)
        ws.cell(row=i, column=8, value=str(linea.cuenta))
        ws.cell(row=i, column=9, value=linea.detalle_item)
        ws.cell(row=i, column=10, value=linea.tipo)
        ws.cell(row=i, column=11, value=linea.importe).number_format = FORMATO_IMPORTE
        ws.cell(row=i, column=12, value=linea.cod_ejercicio)
        ws.cell(row=i, column=13, value=linea.codigo_actividad)
        ws.cell(row=i, column=14, value=linea.nota)

    for j, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ancho

    wb.save(path)
