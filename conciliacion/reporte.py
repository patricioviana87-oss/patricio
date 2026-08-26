"""Genera el Excel de salida de la conciliación (Resumen, Detalle, Por Proveedor, Revisar)."""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FUENTE = "Arial"
COLOR_HEADER = "1F4E78"
COLOR_CONCILIADO = "C6EFCE"
COLOR_PENDIENTE = "FFEB9C"
COLOR_REVISAR = "FFC7CE"

MONEY_FMT = '$#,##0.00;[RED]-$#,##0.00'
DATE_FMT = "DD/MM/YYYY"

ESTADO_FILL = {
    "CONCILIADO": PatternFill("solid", fgColor=COLOR_CONCILIADO),
    "PENDIENTE": PatternFill("solid", fgColor=COLOR_PENDIENTE),
    "ANULADO": PatternFill("solid", fgColor="D9D9D9"),
    "SIN INFORME DE PAGOS": PatternFill("solid", fgColor=COLOR_REVISAR),
    "SIN MOVIMIENTO EN CUENTA": PatternFill("solid", fgColor=COLOR_REVISAR),
}

THIN = Side(style="thin", color="B7B7B7")
BORDE = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = Font(name=FUENTE, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE


def _autosize(ws: Worksheet, n_cols: int, min_width=10, max_width=45) -> None:
    for j in range(1, n_cols + 1):
        col = get_column_letter(j)
        best = min_width
        for cell in ws[col]:
            if cell.value is not None:
                best = max(best, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[col].width = best


def _write_df(ws: Worksheet, df: pd.DataFrame, start_row: int, money_cols: list[str], date_cols: list[str]) -> int:
    headers = list(df.columns)
    _header_row(ws, start_row, headers)
    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        for j, col in enumerate(headers, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name=FUENTE, size=10)
            cell.border = BORDE
            if col in money_cols and val is not None:
                cell.number_format = MONEY_FMT
            if col in date_cols and val is not None:
                cell.number_format = DATE_FMT
        if "ESTADO" in headers:
            estado = row["ESTADO"]
            fill = ESTADO_FILL.get(estado)
            if fill:
                for j in range(1, len(headers) + 1):
                    ws.cell(row=r, column=j).fill = fill
    last_row = start_row + len(df)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{last_row}"
    return last_row


def escribir_reporte(resultado: dict, path_salida: str) -> None:
    wb = Workbook()

    ws_r = wb.active
    ws_r.title = "Resumen"
    r = resultado["resumen"]
    ws_r["A1"] = "Conciliación de cheques / pagos pendientes"
    ws_r["A1"].font = Font(name=FUENTE, bold=True, size=14)

    filas = [
        ("Egresos analizados (en la Query)", r["total_egresos"]),
        ("Conciliados", r["conciliados"]),
        ("Pendientes", r["pendientes"]),
        ("Anulados", r["anulados"]),
        ("Sin informe de pagos (revisar)", r["sin_informe"]),
        ("Pagos del Informe fuera de este período (informativo)", r["pagos_fuera_de_periodo"]),
        (None, None),
        ("Saldo pendiente (según Informe de Pagos)", r["saldo_pendiente"]),
        ("Movimiento neto de la Query cargada", r["total_mayor_cuenta"]),
        ("Diferencia", r["diferencia_control"]),
    ]
    start = 3
    for i, (label, value) in enumerate(filas):
        row = start + i
        if label is None:
            continue
        ws_r.cell(row=row, column=1, value=label).font = Font(name=FUENTE, size=11)
        c = ws_r.cell(row=row, column=2, value=value)
        c.font = Font(name=FUENTE, size=11, bold=label.startswith(("Saldo", "Total", "Diferencia")))
        if isinstance(value, float):
            c.number_format = MONEY_FMT
    ws_r.column_dimensions["A"].width = 42
    ws_r.column_dimensions["B"].width = 20

    nota = (
        "Nota: 'Saldo pendiente' y 'Movimiento neto de la Query' solo van a coincidir si la Query "
        "que cargaste tiene el historico COMPLETO de la cuenta (todos los meses acumulados), como en "
        "el proceso manual. Si cargaste un solo mes, la diferencia es esperable (hay conciliaciones "
        "de ese mes que cancelan egresos de meses anteriores, y egresos de este mes que se van a "
        "conciliar recien el mes que viene). Para un control exacto de saldo, complementar con el "
        "saldo contable real de la cuenta. Revisar la hoja 'Revisar' para los egresos sin cruce contra "
        "el Informe de Pagos o con importe distinto."
    )
    ws_r.cell(row=start + len(filas) + 1, column=1, value=nota).font = Font(name=FUENTE, italic=True, size=9)
    ws_r.merge_cells(start_row=start + len(filas) + 1, start_column=1, end_row=start + len(filas) + 1, end_column=6)

    ws_d = wb.create_sheet("Detalle")
    _write_df(
        ws_d,
        resultado["detalle"],
        start_row=1,
        money_cols=["MONTO", "DIFERENCIA_IMPORTE"],
        date_cols=["FECHA", "FECHA_CONCILIACION"],
    )
    _autosize(ws_d, len(resultado["detalle"].columns))

    ws_p = wb.create_sheet("Por Proveedor")
    pp = resultado["por_proveedor"].rename(columns={"PROVEEDOR": "PROVEEDOR", "MONTO": "MONTO PENDIENTE"})
    last = _write_df(ws_p, pp, start_row=1, money_cols=["MONTO PENDIENTE"], date_cols=[])
    total_row = last + 2
    ws_p.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FUENTE, bold=True)
    col_letter = get_column_letter(list(pp.columns).index("MONTO PENDIENTE") + 1)
    if len(pp) > 0:
        total_value = f"=SUM({col_letter}2:{col_letter}{last})"
    else:
        total_value = 0
    total_cell = ws_p.cell(row=total_row, column=2, value=total_value)
    total_cell.font = Font(name=FUENTE, bold=True)
    total_cell.number_format = MONEY_FMT
    _autosize(ws_p, len(pp.columns))

    ws_x = wb.create_sheet("Revisar")
    if len(resultado["revisar"]) > 0:
        _write_df(
            ws_x,
            resultado["revisar"],
            start_row=1,
            money_cols=["MONTO", "DIFERENCIA_IMPORTE"],
            date_cols=["FECHA", "FECHA_CONCILIACION"],
        )
        _autosize(ws_x, len(resultado["revisar"].columns))
    else:
        ws_x["A1"] = "Sin diferencias para revisar."
        ws_x["A1"].font = Font(name=FUENTE, italic=True)

    ws_f = wb.create_sheet("Fuera de Periodo")
    fuera = resultado["fuera_de_periodo"]
    if len(fuera) > 0:
        ws_f["A1"] = "Pagos del Informe que no corresponden a ningún egreso de la Query cargada (informativo)"
        ws_f["A1"].font = Font(name=FUENTE, italic=True, size=9)
        ws_f.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        _write_df(ws_f, fuera, start_row=2, money_cols=["MONTO"], date_cols=["FECHA"])
        _autosize(ws_f, len(fuera.columns))
    else:
        ws_f["A1"] = "No hay pagos fuera de este período."
        ws_f["A1"].font = Font(name=FUENTE, italic=True)

    wb.save(path_salida)
